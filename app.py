import os
import csv
import requests
import pandas as pd
import io
from datetime import datetime
import urllib.parse
import openpyxl
from openpyxl.styles import PatternFill
import streamlit as st
import re

def validate_sic_input(input_str):
    """Validate manually entered SIC codes"""
    if not input_str.strip():
        return False, "Input cannot be empty"
    
    # Check for invalid characters (only numbers, commas, and spaces allowed)
    if not re.match(r'^[\d\s,]+$', input_str):
        return False, "Only numbers and commas are allowed"
    
    # Split and check individual codes
    codes = [code.strip() for code in input_str.split(',') if code.strip()]
    if not codes:
        return False, "No valid SIC codes found"
    
    for code in codes:
        if not code.isdigit():
            return False, f"'{code}' is not a valid number"
        if len(code) > 5:  # SIC codes are typically 4-5 digits
            return False, f"'{code}' is too long (max 5 digits)"
    
    return True, codes

def read_keywords_from_file(file):
    """Read keywords from uploaded file"""
    try:
        df = pd.read_csv(file)
        # Assuming keywords are in the first column
        keywords = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
        
        # Validate the codes from file
        valid_codes = []
        invalid_codes = []
        for code in keywords:
            if code.isdigit() and len(code) <= 5:
                valid_codes.append(code)
            else:
                invalid_codes.append(code)
        
        if invalid_codes:
            st.warning(f"Found {len(invalid_codes)} invalid SIC codes in file (ignored): {', '.join(invalid_codes)}")
        
        return valid_codes
    except Exception as e:
        st.error(f"Error reading parameters file: {e}")
        return []

def download_csv_for_keyword(keyword):
    """Download CSV for a specific keyword"""
    try:
        encoded_keyword = urllib.parse.quote(keyword)
        url = f"https://find-and-update.company-information.service.gov.uk/advanced-search/download?companyNameIncludes=&companyNameExcludes=&registeredOfficeAddress=LE5&incorporationFromDay=&incorporationFromMonth=&incorporationFromYear=&incorporationToDay=&incorporationToMonth=&incorporationToYear=&sicCodes={encoded_keyword}&dissolvedFromDay=&dissolvedFromMonth=&dissolvedFromYear=&dissolvedToDay=&dissolvedToMonth=&dissolvedToYear="
        response = requests.get(url)
        response.raise_for_status()
        df = pd.read_csv(io.StringIO(response.text))
        return df
    except Exception as e:
        st.warning(f"Error downloading CSV for keyword {keyword}: {e}")
        return None

def process_dissolution_date(date_str):
    """Process dissolution date string"""
    if pd.isna(date_str) or date_str == '':
        return None
    try:
        return datetime.strptime(date_str, '%d/%m/%Y')
    except:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d')
        except:
            return None

def filter_and_append_data(input_df):
    """Filter data based on dissolution date"""
    filtered_df = input_df[
        (input_df['dissolution_date'].apply(process_dissolution_date).isna()) | 
        (input_df['dissolution_date'].apply(process_dissolution_date) > datetime(2019, 1, 1))
    ]
    return filtered_df

def process_company_stats(data_df):
    """Generate statistics about companies"""
    stats = {
        'Active Companies': len(data_df[data_df['company_status'] == 'Active']),
        'Dissolved Companies': len(data_df[data_df['company_status'] == 'Dissolved']),
        'Companies in Liquidation': len(data_df[data_df['company_status'] == 'Liquidation'])
    }
    
    data_df['dissolution_year'] = data_df['dissolution_date'].apply(
        lambda x: process_dissolution_date(x).year if process_dissolution_date(x) else None
    )
    
    dissolution_by_year = data_df[data_df['dissolution_year'].notna()]['dissolution_year'].value_counts().to_dict()
    stats['Dissolution by Year'] = dissolution_by_year
    
    return stats

def highlight_dissolved_rows(sheet, max_col):
    """Highlight dissolved companies in a worksheet"""
    red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=max_col):
        if row[max_col - 1].value == 'Dissolved':
            for cell in row:
                cell.fill = red_fill

def main():
    st.title("Company Data Processor")
    st.write("Search for companies by SIC codes. You can either upload a CSV file or enter codes manually.")
    
    # Input method selection
    input_method = st.radio(
        "Select input method:",
        ("Upload CSV file", "Enter SIC codes manually"),
        horizontal=True
    )
    
    keywords = []
    
    if input_method == "Upload CSV file":
        uploaded_file = st.file_uploader("Upload CSV file with SIC codes (one per line or comma-separated)", type=['csv'])
        if uploaded_file is not None:
            keywords = read_keywords_from_file(uploaded_file)
    else:
        manual_input = st.text_area(
            "Enter SIC codes (comma-separated):",
            placeholder="e.g., 62012, 62020, 58210",
            help="Enter 4-5 digit SIC codes separated by commas"
        )
        
        if st.button("Validate and Process Codes"):
            is_valid, result = validate_sic_input(manual_input)
            if is_valid:
                keywords = result
                st.success(f"Validated {len(keywords)} SIC codes")
            else:
                st.error(f"Validation error: {result}")
    
    if keywords:
        with st.spinner('Processing...'):
            # Accumulated data
            all_data = pd.DataFrame()
            active_company_addresses = pd.DataFrame(columns=['company_name', 'registered_office_address'])
            processed_sheets = {}  # To store processed data for each SIC code
            
            # Process each keyword
            progress_bar = st.progress(0)
            for idx, keyword in enumerate(keywords):
                progress_bar.progress((idx + 1) / len(keywords))
                
                # Download CSV
                keyword_df = download_csv_for_keyword(keyword)
                
                if keyword_df is not None:
                    # Filter data
                    filtered_df = filter_and_append_data(keyword_df)
                    
                    # Store processed data for this SIC code
                    processed_sheets[keyword] = filtered_df
                    
                    # Append to accumulated data
                    all_data = pd.concat([all_data, filtered_df], ignore_index=True)
                    
                    # Collect active company addresses
                    active_addresses = filtered_df[filtered_df['company_status'] == 'Active'][['company_name', 'registered_office_address']]
                    active_company_addresses = pd.concat([active_company_addresses, active_addresses], ignore_index=True)
            
            if not all_data.empty:
                # Remove duplicates
                all_data.drop_duplicates(inplace=True)
                active_company_addresses.drop_duplicates(inplace=True)
                
                # Generate stats
                stats = process_company_stats(all_data)
                
                # Create Excel file in memory
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Write individual sheets for each processed SIC code
                    for sic_code, df in processed_sheets.items():
                        # Clean sheet name to be valid Excel sheet name
                        sheet_name = f"SIC_{sic_code}"[:31]  # Excel sheet names max 31 chars
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Highlight dissolved companies in each sheet
                        workbook = writer.book
                        sheet = workbook[sheet_name]
                        highlight_dissolved_rows(sheet, df.shape[1])
                    
                    # Write master data sheet
                    all_data.to_excel(writer, sheet_name='Master_Data', index=False)
                    highlight_dissolved_rows(writer.book['Master_Data'], all_data.shape[1])
                    
                    # Write stats and addresses sheets
                    pd.DataFrame.from_dict(stats, orient='index').to_excel(writer, sheet_name='Stats')
                    active_company_addresses.to_excel(writer, sheet_name='Active_Addresses', index=False)
                
                output.seek(0)
                
                # Show results
                st.success("Processing complete!")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.download_button(
                        label="Download Excel Report",
                        data=output,
                        file_name="Company_Data.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with col2:
                    st.download_button(
                        label="Download Active Addresses (CSV)",
                        data=active_company_addresses.to_csv(index=False).encode('utf-8'),
                        file_name="Active_Addresses.csv",
                        mime="text/csv"
                    )
                
                # Show previews
                st.subheader("Preview of Master Data")
                st.dataframe(all_data.head())
                
                st.subheader("Preview of Active Company Addresses")
                st.dataframe(active_company_addresses.head())
                
                st.subheader("Statistics")
                st.write(stats)
            else:
                st.warning("No valid company data found for the provided SIC codes")

if __name__ == "__main__":
    main()
